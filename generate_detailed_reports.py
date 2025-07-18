import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

def load_json_file(filepath):
    """Load JSON file if it exists, return None otherwise"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return None

def load_markdown_file(filepath):
    """Load markdown file if it exists, return None otherwise"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return None

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Handle arrays by converting to string representation
            items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)

def create_section_header(ws, row, title, start_col=1, end_col=10):
    """Create a formatted section header"""
    # Merge cells for the header
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    
    # Set the title
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    return row + 1

def add_dataframe_to_worksheet(ws, df, start_row, title=""):
    """Add a dataframe to worksheet with formatting"""
    if title:
        start_row = create_section_header(ws, start_row, title, end_col=len(df.columns))
        start_row += 1
    
    # Add headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # Add data
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return row_idx + 3  # Return next available row with spacing

def create_classification_section(results_data, filename):
    """Create classification section dataframe"""
    if not results_data or 'classification_result' not in results_data:
        return pd.DataFrame()
    
    classification = results_data['classification_result']
    flattened = flatten_dict(classification)
    
    # Create single row dataframe
    data = {'file': filename}
    data.update(flattened)
    data['QA'] = ''
    data['note'] = ''
    
    return pd.DataFrame([data])

def create_extraction_section(results_data, filename):
    """Create extraction section dataframe"""
    if not results_data or 'extraction_result' not in results_data:
        return pd.DataFrame()
    
    extraction = results_data['extraction_result']
    rows = []
    
    for key, value in extraction.items():
        if isinstance(value, dict):
            # Handle nested objects
            for sub_key, sub_value in value.items():
                rows.append({
                    'file_name': filename,
                    'field_key': f"{key}_{sub_key}",
                    'field_value': str(sub_value) if sub_value is not None else '',
                    'QA': '',
                    'note': ''
                })
        elif isinstance(value, list):
            # Handle arrays
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    for sub_key, sub_value in item.items():
                        rows.append({
                            'file_name': filename,
                            'field_key': f"{key}[{i}]_{sub_key}",
                            'field_value': str(sub_value) if sub_value is not None else '',
                            'QA': '',
                            'note': ''
                        })
                else:
                    rows.append({
                        'file_name': filename,
                        'field_key': f"{key}[{i}]",
                        'field_value': str(item) if item is not None else '',
                        'QA': '',
                        'note': ''
                    })
        else:
            rows.append({
                'file_name': filename,
                'field_key': key,
                'field_value': str(value) if value is not None else '',
                'QA': '',
                'note': ''
            })
    
    return pd.DataFrame(rows)

def create_issues_section(results_data, filename):
    """Create issues section dataframe"""
    if not results_data or 'compliance_result' not in results_data:
        return pd.DataFrame()
    
    compliance = results_data['compliance_result']
    validation = compliance.get('validation_result', {})
    issues = validation.get('issues', [])
    
    rows = []
    for i, issue in enumerate(issues, 1):
        rows.append({
            'index': i,
            'issue_type': issue.get('issue_type', ''),
            'field': issue.get('field', ''),
            'description': issue.get('description', ''),
            'recommendation': issue.get('recommendation', ''),
            'knowledge_base_reference': issue.get('knowledge_base_reference', ''),
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_validation_overall_section(validation_data, filename):
    """Create validation overall assessment section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    overall = validation_data['validation_report'].get('overall_assessment', {})
    data = {'file': filename}
    data.update(overall)
    data['QA'] = ''
    data['note'] = ''
    
    return pd.DataFrame([data])

def create_validation_dimensions_section(validation_data, filename):
    """Create validation dimension details section"""
    if not validation_data or 'detailed_analysis' not in validation_data:
        return pd.DataFrame()
    
    dimension_details = validation_data['detailed_analysis'].get('dimension_details', {})
    rows = []
    
    for dimension, details in dimension_details.items():
        row = {
            'file': filename,
            'dimension': dimension,
            'confidence_score': details.get('confidence_score', ''),
            'reliability_level': details.get('reliability_level', ''),
            'summary': details.get('summary', ''),
            'total_issues': details.get('total_issues', ''),
            'QA': '',
            'note': ''
        }
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_validation_summary_section(validation_data, filename):
    """Create dimensional analysis summary section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    dimensional_summary = validation_data['validation_report'].get('dimensional_analysis_summary', {})
    rows = []
    
    for dimension, details in dimensional_summary.items():
        row = {
            'file': filename,
            'dimension': dimension,
            'confidence': details.get('confidence', ''),
            'reliability': details.get('reliability', ''),
            'issues_count': details.get('issues_count', ''),
            'QA': '',
            'note': ''
        }
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_validation_critical_section(validation_data, filename):
    """Create critical issues summary section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    critical_issues = validation_data['validation_report'].get('critical_issues_summary', {}).get('issues', [])
    rows = []
    
    for i, issue in enumerate(critical_issues, 1):
        rows.append({
            'file': filename,
            'index': i,
            'critical_issue': issue,
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_markdown_section(markdown_content, filename):
    """Create markdown section dataframe"""
    if not markdown_content:
        return pd.DataFrame()
    
    return pd.DataFrame([{
        'file_name': filename,
        'markdown_content': markdown_content,
        'QA': '',
        'note': ''
    }])

def create_image_quality_overall_section(quality_data, filename):
    """Create image quality overall assessment section"""
    if not quality_data:
        return pd.DataFrame()
    
    overall = quality_data.get('overall_assessment', {})
    image_type = quality_data.get('image_type_detection', {})
    
    data = {
        'file': filename,
        'overall_score': overall.get('score', ''),
        'quality_level': overall.get('level', ''),
        'quality_passed': overall.get('pass_fail', ''),
        'processing_time_seconds': quality_data.get('processing_time_seconds', ''),
        'image_type': image_type.get('image_subtype', ''),
        'is_digital_screenshot': image_type.get('is_digital_screenshot', ''),
        'confidence': image_type.get('confidence', ''),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_image_quality_detailed_section(quality_data, filename):
    """Create image quality detailed metrics section"""
    if not quality_data or 'detailed_results' not in quality_data:
        return pd.DataFrame()
    
    detailed = quality_data['detailed_results']
    rows = []
    
    # Resolution metrics
    if 'resolution' in detailed:
        res = detailed['resolution']
        dimensions = res.get('dimensions', {})
        dpi = res.get('dpi', {})
        quality = res.get('quality', {})
        
        key_metrics = f"Dimensions: {dimensions.get('width', '')}x{dimensions.get('height', '')}, DPI: {dpi.get('average', '')}, Megapixels: {dimensions.get('megapixels', '')}"
        recommendations = '; '.join(res.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Resolution',
            'score': quality.get('score', ''),
            'level': quality.get('level', ''),
            'meets_requirements': quality.get('meets_ocr_requirements', ''),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Blur metrics
    if 'blur' in detailed:
        blur = detailed['blur']
        metrics = blur.get('metrics', {})
        motion_blur = blur.get('motion_blur', {})
        
        key_metrics = f"Laplacian Variance: {metrics.get('laplacian_variance', '')}, Motion Blur: {motion_blur.get('detected', '')}, Direction: {motion_blur.get('direction', '')}"
        recommendations = '; '.join(blur.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Blur',
            'score': metrics.get('blur_score', ''),
            'level': metrics.get('blur_level', ''),
            'meets_requirements': not metrics.get('is_blurry', True),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Glare metrics
    if 'glare' in detailed:
        glare = detailed['glare']
        exposure = glare.get('exposure_metrics', {})
        analysis = glare.get('glare_analysis', {})
        
        key_metrics = f"Mean Brightness: {exposure.get('mean_brightness', '')}, Overexposed: {exposure.get('overexposed_percent', '')}%, Glare Spots: {analysis.get('num_glare_spots', '')}"
        recommendations = '; '.join(glare.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Glare',
            'score': analysis.get('glare_score', ''),
            'level': analysis.get('glare_level', ''),
            'meets_requirements': not exposure.get('is_overexposed', False),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Completeness metrics
    if 'completeness' in detailed:
        completeness = detailed['completeness']
        edge_analysis = completeness.get('edge_analysis', {})
        corner_analysis = completeness.get('corner_analysis', {})
        
        key_metrics = f"Boundary Detected: {completeness.get('boundary_detected', '')}, Edge Coverage: {edge_analysis.get('edge_coverage', '')}%, Visible Corners: {corner_analysis.get('visible_corners', '')}"
        recommendations = '; '.join(completeness.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Completeness',
            'score': completeness.get('completeness_score', ''),
            'level': completeness.get('completeness_level', ''),
            'meets_requirements': completeness.get('boundary_detected', ''),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Damage metrics
    if 'damage' in detailed:
        damage = detailed['damage']
        stain = damage.get('stain_analysis', {})
        tear = damage.get('tear_analysis', {})
        fold = damage.get('fold_analysis', {})
        
        key_metrics = f"Damage Types: {', '.join(damage.get('damage_types', []))}, Stains: {stain.get('count', 0)}, Tears: {tear.get('count', 0)}, Folds: {fold.get('count', 0)}"
        recommendations = '; '.join(damage.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Damage',
            'score': damage.get('damage_score', ''),
            'level': damage.get('damage_level', ''),
            'meets_requirements': len(damage.get('damage_types', [])) == 0,
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_image_quality_breakdown_section(quality_data, filename):
    """Create image quality score breakdown section"""
    if not quality_data or 'score_breakdown' not in quality_data:
        return pd.DataFrame()
    
    breakdown = quality_data['score_breakdown']
    rows = []
    
    for component, details in breakdown.items():
        rows.append({
            'file': filename,
            'quality_component': component.title(),
            'individual_score': details.get('score', ''),
            'weight': details.get('weight', ''),
            'contribution': details.get('contribution', ''),
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_citation_section(citation_data, filename):
    """Create citation section dataframe with field-level citations"""
    if not citation_data or 'citations' not in citation_data:
        return pd.DataFrame()
    
    citations = citation_data['citations']
    rows = []
    
    for field_name, citation_info in citations.items():
        field_citation = citation_info.get('field_citation', {})
        value_citation = citation_info.get('value_citation', {})
        
        # Handle case where value_citation might be None
        if value_citation is None:
            value_citation = {}
        
        row = {
            'file_name': filename,
            'field_name': field_name,
            'field_source_text': field_citation.get('source_text', ''),
            'field_confidence': field_citation.get('confidence', ''),
            'field_source_location': field_citation.get('source_location', ''),
            'field_match_type': field_citation.get('match_type', ''),
            'value_source_text': value_citation.get('source_text', ''),
            'value_confidence': value_citation.get('confidence', ''),
            'value_source_location': value_citation.get('source_location', ''),
            'value_match_type': value_citation.get('match_type', ''),
            'context': field_citation.get('context', ''),
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_citation_metadata_section(citation_data, filename):
    """Create citation metadata summary section"""
    if not citation_data or 'metadata' not in citation_data:
        return pd.DataFrame()
    
    metadata = citation_data['metadata']
    
    data = {
        'file_name': filename,
        'total_fields_analyzed': metadata.get('total_fields_analyzed', 0),
        'fields_with_field_citations': metadata.get('fields_with_field_citations', 0),
        'fields_with_value_citations': metadata.get('fields_with_value_citations', 0),
        'average_confidence': metadata.get('average_confidence', 0),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_llm_quality_overall_section(llm_quality_data, filename):
    """Create LLM quality overall assessment section"""
    if not llm_quality_data:
        return pd.DataFrame()
    
    data = {
        'file_name': filename,
        'assessment_method': llm_quality_data.get('assessment_method', ''),
        'model_used': llm_quality_data.get('model_used', ''),
        'overall_quality_score': llm_quality_data.get('overall_quality_score', ''),
        'suitable_for_extraction': llm_quality_data.get('suitable_for_extraction', ''),
        'timestamp': llm_quality_data.get('timestamp', ''),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_llm_quality_detailed_section(llm_quality_data, filename):
    """Create LLM quality detailed assessment section"""
    if not llm_quality_data:
        return pd.DataFrame()
    
    # Define the quality aspects to extract
    quality_aspects = [
        'blur_detection',
        'contrast_assessment',
        'glare_identification',
        'water_stains',
        'tears_or_folds',
        'cut_off_detection',
        'missing_sections',
        'obstructions'
    ]
    
    rows = []
    
    for aspect in quality_aspects:
        if aspect in llm_quality_data:
            aspect_data = llm_quality_data[aspect]
            
            row = {
                'file_name': filename,
                'quality_aspect': aspect.replace('_', ' ').title(),
                'detected': aspect_data.get('detected', ''),
                'severity_level': aspect_data.get('severity_level', ''),
                'confidence_score': aspect_data.get('confidence_score', ''),
                'quantitative_measure': aspect_data.get('quantitative_measure', ''),
                'description': aspect_data.get('description', ''),
                'recommendation': aspect_data.get('recommendation', ''),
                'QA': '',
                'note': ''
            }
            
            rows.append(row)
    
    return pd.DataFrame(rows)

def auto_adjust_column_widths(ws):
    """Auto-adjust column widths for a worksheet"""
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_worksheet_for_file(wb, filename):
    """Create a worksheet for a single file with all sections"""
    print(f"Processing {filename}...")
    
    # Load data from all sources
    results_data = load_json_file(f'results/{filename}.json')
    validation_data = load_json_file(f'validation_results/{filename}_validation.json')
    quality_data = load_json_file(f'quality_reports/{filename}_quality.json')
    markdown_content = load_markdown_file(f'llamaparse_output/{filename}.md')
    citation_data = load_json_file(f'citation_folder/{filename}_citation.json')
    llm_quality_data = load_json_file(f'llm_quality_reports/llm_quality_{filename}.json')
    
    # Create worksheet
    ws = wb.create_sheet(title=filename)
    current_row = 1
    
    # 1. Classification Section
    classification_df = create_classification_section(results_data, filename)
    if not classification_df.empty:
        current_row = add_dataframe_to_worksheet(ws, classification_df, current_row, "CLASSIFICATION SECTION")
    
    # 2. Extraction Results Section
    extraction_df = create_extraction_section(results_data, filename)
    if not extraction_df.empty:
        current_row = add_dataframe_to_worksheet(ws, extraction_df, current_row, "EXTRACTION RESULTS SECTION")
    
    # 3. Issues Section
    issues_df = create_issues_section(results_data, filename)
    if not issues_df.empty:
        current_row = add_dataframe_to_worksheet(ws, issues_df, current_row, "ISSUES SECTION")
    
    # 4. Validation - Overall Assessment
    validation_overall_df = create_validation_overall_section(validation_data, filename)
    if not validation_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_overall_df, current_row, "VALIDATION - OVERALL ASSESSMENT")
    
    # 5. Validation - Dimension Details
    validation_dimensions_df = create_validation_dimensions_section(validation_data, filename)
    if not validation_dimensions_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_dimensions_df, current_row, "VALIDATION - DIMENSION DETAILS")
    
    # 6. Validation - Dimensional Analysis Summary
    validation_summary_df = create_validation_summary_section(validation_data, filename)
    if not validation_summary_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_summary_df, current_row, "VALIDATION - DIMENSIONAL ANALYSIS SUMMARY")
    
    # 7. Validation - Critical Issues
    validation_critical_df = create_validation_critical_section(validation_data, filename)
    if not validation_critical_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_critical_df, current_row, "VALIDATION - CRITICAL ISSUES")
    
    # 8. Markdown Section
    markdown_df = create_markdown_section(markdown_content, filename)
    if not markdown_df.empty:
        current_row = add_dataframe_to_worksheet(ws, markdown_df, current_row, "MARKDOWN SECTION")
    
    # 9. Citation Section - Field-Level Citations
    citation_df = create_citation_section(citation_data, filename)
    if not citation_df.empty:
        current_row = add_dataframe_to_worksheet(ws, citation_df, current_row, "CITATION SECTION - FIELD-LEVEL CITATIONS")
    
    # 10. Citation Section - Metadata Summary
    citation_metadata_df = create_citation_metadata_section(citation_data, filename)
    if not citation_metadata_df.empty:
        current_row = add_dataframe_to_worksheet(ws, citation_metadata_df, current_row, "CITATION SECTION - METADATA SUMMARY")
    
    # 11. Image Quality - Overall Assessment
    quality_overall_df = create_image_quality_overall_section(quality_data, filename)
    if not quality_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, quality_overall_df, current_row, "IMAGE QUALITY - OVERALL ASSESSMENT")
    
    # 12. Image Quality - Detailed Metrics
    quality_detailed_df = create_image_quality_detailed_section(quality_data, filename)
    if not quality_detailed_df.empty:
        current_row = add_dataframe_to_worksheet(ws, quality_detailed_df, current_row, "IMAGE QUALITY - DETAILED METRICS")
    
    # 13. Image Quality - Score Breakdown
    quality_breakdown_df = create_image_quality_breakdown_section(quality_data, filename)
    if not quality_breakdown_df.empty:
        current_row = add_dataframe_to_worksheet(ws, quality_breakdown_df, current_row, "IMAGE QUALITY - SCORE BREAKDOWN")
    
    # 14. LLM Quality - Overall Assessment
    llm_quality_overall_df = create_llm_quality_overall_section(llm_quality_data, filename)
    if not llm_quality_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_quality_overall_df, current_row, "LLM QUALITY - OVERALL ASSESSMENT")
    
    # 15. LLM Quality - Detailed Assessment
    llm_quality_detailed_df = create_llm_quality_detailed_section(llm_quality_data, filename)
    if not llm_quality_detailed_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_quality_detailed_df, current_row, "LLM QUALITY - DETAILED ASSESSMENT")
    
    # Auto-adjust column widths
    auto_adjust_column_widths(ws)
    
    print(f"Added worksheet: {filename}")

def get_file_list():
    """Get list of all unique file names across all directories"""
    files = set()
    
    # Check each directory for files
    directories = ['results', 'validation_results', 'quality_reports', 'dataset', 'citation_folder', 'llm_quality_reports']
    
    for directory in directories:
        if os.path.exists(directory):
            for filename in os.listdir(directory):
                if filename.endswith('.json'):
                    base_name = filename.replace('.json', '')
                    
                    # Handle different naming conventions
                    if base_name.endswith('_validation'):
                        base_name = base_name.replace('_validation', '')
                    elif base_name.endswith('_quality'):
                        base_name = base_name.replace('_quality', '')
                    elif base_name.endswith('_citation'):
                        base_name = base_name.replace('_citation', '')
                    elif base_name.startswith('llm_quality_'):
                        base_name = base_name.replace('llm_quality_', '')
                    
                    # Filter out summary_report as it's not an individual file
                    if base_name != 'summary_report':
                        files.add(base_name)
    
    return sorted(list(files))

def extract_results_data_summary(data):
    """Extract data from results folder JSON for summary"""
    if not data:
        return {}
    
    # Extract classification data
    classification = data.get('classification_result', {})
    
    # Extract compliance issues count
    compliance = data.get('compliance_result', {})
    validation = compliance.get('validation_result', {})
    issues_count = validation.get('issues_count', 0)
    
    return {
        'issues_count': issues_count,
        'is_expense': classification.get('is_expense', False),
        'language': classification.get('language', ''),
        'language_confidence': classification.get('language_confidence', 0),
        'classification_confidence': classification.get('classification_confidence', 0)
    }

def extract_validation_data_summary(data):
    """Extract data from validation_results folder JSON for summary"""
    if not data:
        return {}
    
    validation_report = data.get('validation_report', {})
    overall = validation_report.get('overall_assessment', {})
    dimensional = validation_report.get('dimensional_analysis_summary', {})
    
    return {
        'confidence_score': overall.get('confidence_score', 0),
        'is_reliable': overall.get('is_reliable', False),
        'hallucination_detection': dimensional.get('hallucination_detection', {}).get('confidence', 0),
        'compliance_accuracy': dimensional.get('compliance_accuracy', {}).get('confidence', 0),
        'factual_grounding': dimensional.get('factual_grounding', {}).get('confidence', 0),
        'knowledge_base_adherence': dimensional.get('knowledge_base_adherence', {}).get('confidence', 0)
    }

def extract_quality_data_summary(data):
    """Extract data from quality_reports folder JSON for summary"""
    if not data:
        return {}
    
    overall = data.get('overall_assessment', {})
    detailed = data.get('detailed_results', {})
    
    return {
        'quality_score': overall.get('score', 0),
        'quality_level': overall.get('level', ''),
        'quality_passed': overall.get('pass_fail', False),
        'resolution_score': detailed.get('resolution', {}).get('quality', {}).get('score', 0),
        'blur_score': detailed.get('blur', {}).get('metrics', {}).get('blur_score', 0),
        'glare_score': detailed.get('glare', {}).get('glare_analysis', {}).get('glare_score', 0),
        'completeness_score': detailed.get('completeness', {}).get('completeness_score', 0),
        'damage_score': detailed.get('damage', {}).get('damage_score', 0)
    }

def extract_llm_quality_data_summary(data):
    """Extract data from llm_quality_reports folder JSON for summary"""
    if not data:
        return {}
    
    return {
        'llm_suitable_for_extraction': data.get('suitable_for_extraction', False),
        'llm_model_used': data.get('model_used', ''),
        'llm_blur_detected': data.get('blur_detection', {}).get('detected', False),
        'llm_blur_severity': data.get('blur_detection', {}).get('severity_level', ''),
        'llm_glare_detected': data.get('glare_identification', {}).get('detected', False),
        'llm_glare_severity': data.get('glare_identification', {}).get('severity_level', ''),
        'llm_cut_off_detected': data.get('cut_off_detection', {}).get('detected', False),
        'llm_missing_sections': data.get('missing_sections', {}).get('detected', False)
    }

def create_summary_worksheet(wb, file_list):
    """Create summary worksheet as first tab"""
    print("Creating summary worksheet...")
    
    # Initialize data list
    report_data = []
    
    for file_base in file_list:
        # Initialize row data
        row_data = {'file': file_base}
        
        # Load data from each source
        results_data = load_json_file(f'results/{file_base}.json')
        validation_data = load_json_file(f'validation_results/{file_base}_validation.json')
        quality_data = load_json_file(f'quality_reports/{file_base}_quality.json')
        dataset_data = load_json_file(f'dataset/{file_base}.json')
        llm_quality_data = load_json_file(f'llm_quality_reports/llm_quality_{file_base}.json')
        
        # Extract data from each source
        row_data.update(extract_results_data_summary(results_data))
        row_data.update(extract_validation_data_summary(validation_data))
        row_data.update(extract_quality_data_summary(quality_data))
        row_data.update(extract_llm_quality_data_summary(llm_quality_data))
        
        # Add dataset info if available
        if dataset_data:
            row_data['country'] = dataset_data.get('country', '')
            row_data['icp'] = dataset_data.get('icp', '')
        
        report_data.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(report_data)
    
    # Define column order with QA columns interspersed
    columns_order = ['file']
    data_columns = [
        'issues_count',
        'is_expense',
        'language',
        'language_confidence',
        'classification_confidence',
        'confidence_score',
        'is_reliable',
        'hallucination_detection',
        'compliance_accuracy',
        'factual_grounding',
        'knowledge_base_adherence',
        'quality_score',
        'quality_level',
        'quality_passed',
        'resolution_score',
        'blur_score',
        'glare_score',
        'completeness_score',
        'damage_score',
        'llm_suitable_for_extraction',
        'llm_model_used',
        'llm_blur_detected',
        'llm_blur_severity',
        'llm_glare_detected',
        'llm_glare_severity',
        'llm_cut_off_detected',
        'llm_missing_sections',
        'country',
        'icp'
    ]
    
    # Add each data column followed by QA column
    for col in data_columns:
        if col in df.columns:
            columns_order.append(col)
            columns_order.append('QA')
    
    # Add QA columns to dataframe
    for col in data_columns:
        if col in df.columns:
            df[f'{col}_QA'] = ''
    
    # Reorder columns (only include columns that exist)
    final_columns = []
    for col in columns_order:
        if col == 'QA':
            # Find the previous data column and add its QA column
            prev_col = final_columns[-1] if final_columns else None
            if prev_col and prev_col != 'file':
                qa_col = f'{prev_col}_QA'
                if qa_col in df.columns:
                    final_columns.append(qa_col)
        elif col in df.columns:
            final_columns.append(col)
    
    df = df[final_columns]
    
    # Fill NaN values with appropriate defaults
    df = df.fillna({
        'issues_count': 0,
        'is_expense': False,
        'language': '',
        'language_confidence': 0,
        'classification_confidence': 0,
        'confidence_score': 0,
        'is_reliable': False,
        'hallucination_detection': 0,
        'compliance_accuracy': 0,
        'factual_grounding': 0,
        'knowledge_base_adherence': 0,
        'quality_score': 0,
        'quality_level': '',
        'quality_passed': False,
        'resolution_score': 0,
        'blur_score': 0,
        'glare_score': 0,
        'completeness_score': 0,
        'damage_score': 0,
        'llm_suitable_for_extraction': False,
        'llm_model_used': '',
        'llm_blur_detected': False,
        'llm_blur_severity': '',
        'llm_glare_detected': False,
        'llm_glare_severity': '',
        'llm_cut_off_detected': False,
        'llm_missing_sections': False,
        'country': '',
        'icp': ''
    })
    
    # Create summary worksheet
    ws = wb.create_sheet(title="Summary", index=0)
    current_row = 1
    
    # Add summary data to worksheet
    current_row = add_dataframe_to_worksheet(ws, df, current_row, "EXPENSE ANALYSIS SUMMARY REPORT")
    
    # Auto-adjust column widths
    auto_adjust_column_widths(ws)
    
    print(f"Added summary worksheet with {len(df)} files")

def generate_multitab_report():
    """Generate multi-tab Excel report with one tab per file"""
    file_list = get_file_list()
    print(f"Found {len(file_list)} files to process: {file_list}")
    
    # Create workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create summary worksheet first
    create_summary_worksheet(wb, file_list)
    
    # Create worksheets for each file
    for filename in file_list:
        try:
            create_worksheet_for_file(wb, filename)
        except Exception as e:
            print(f"Error processing {filename}: {e}")
    
    # Save the workbook
    output_file = 'consolidated_expense_reports.xlsx'
    wb.save(output_file)
    
    print(f"\nSuccessfully generated multi-tab report: {output_file}")
    print(f"Report contains {len(wb.worksheets)} worksheets:")
    for ws in wb.worksheets:
        print(f"  - {ws.title}")
    
    return output_file

if __name__ == "__main__":
    generate_multitab_report()
