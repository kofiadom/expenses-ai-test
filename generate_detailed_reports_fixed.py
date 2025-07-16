import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

def load_json_file(filepath):
    """Load JSON file if it exists, return None otherwise"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return None

def load_markdown_file(filepath):
    """Load markdown file if it exists, return None otherwise"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return None

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Handle arrays by converting to string representation
            items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)

def create_section_header(ws, row, title, start_col=1, end_col=10):
    """Create a formatted section header"""
    # Merge cells for the header
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    
    # Set the title
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    return row + 1

def add_dataframe_to_worksheet(ws, df, start_row, title=""):
    """Add a dataframe to worksheet with formatting"""
    if title:
        start_row = create_section_header(ws, start_row, title, end_col=len(df.columns))
        start_row += 1
    
    # Add headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # Add data
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return row_idx + 3  # Return next available row with spacing

def create_classification_section(results_data, filename):
    """Create classification section dataframe"""
    if not results_data or 'classification_result' not in results_data:
        return pd.DataFrame()
    
    classification = results_data['classification_result']
    flattened = flatten_dict(classification)
    
    # Create single row dataframe
    data = {'file': filename}
    data.update(flattened)
    data['QA'] = ''
    data['note'] = ''
    
    return pd.DataFrame([data])

def create_extraction_section(results_data, filename):
    """Create extraction section dataframe"""
    if not results_data or 'extraction_result' not in results_data:
        return pd.DataFrame()
    
    extraction = results_data['extraction_result']
    rows = []
    
    for key, value in extraction.items():
        if isinstance(value, dict):
            # Handle nested objects
            for sub_key, sub_value in value.items():
                rows.append({
                    'file_name': filename,
                    'field_key': f"{key}_{sub_key}",
                    'field_value': str(sub_value) if sub_value is not None else '',
                    'QA': '',
                    'note': ''
                })
        elif isinstance(value, list):
            # Handle arrays
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    for sub_key, sub_value in item.items():
                        rows.append({
                            'file_name': filename,
                            'field_key': f"{key}[{i}]_{sub_key}",
                            'field_value': str(sub_value) if sub_value is not None else '',
                            'QA': '',
                            'note': ''
                        })
                else:
                    rows.append({
                        'file_name': filename,
                        'field_key': f"{key}[{i}]",
                        'field_value': str(item) if item is not None else '',
                        'QA': '',
                        'note': ''
                    })
        else:
            rows.append({
                'file_name': filename,
                'field_key': key,
                'field_value': str(value) if value is not None else '',
                'QA': '',
                'note': ''
            })
    
    return pd.DataFrame(rows)

def create_issues_section(results_data, filename):
    """Create issues section dataframe"""
    if not results_data or 'compliance_result' not in results_data:
        return pd.DataFrame()
    
    compliance = results_data['compliance_result']
    validation = compliance.get('validation_result', {})
    issues = validation.get('issues', [])
    
    rows = []
    for i, issue in enumerate(issues, 1):
        rows.append({
            'index': i,
            'issue_type': issue.get('issue_type', ''),
            'field': issue.get('field', ''),
            'description': issue.get('description', ''),
            'recommendation': issue.get('recommendation', ''),
            'knowledge_base_reference': issue.get('knowledge_base_reference', ''),
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_validation_overall_section(validation_data, filename):
    """Create validation overall assessment section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    overall = validation_data['validation_report'].get('overall_assessment', {})
    data = {'file': filename}
    data.update(overall)
    data['QA'] = ''
    data['note'] = ''
    
    return pd.DataFrame([data])

def create_validation_dimensions_section(validation_data, filename):
    """Create validation dimension details section"""
    if not validation_data or 'detailed_analysis' not in validation_data:
        return pd.DataFrame()
    
    dimension_details = validation_data['detailed_analysis'].get('dimension_details', {})
    rows = []
    
    for dimension, details in dimension_details.items():
        row = {
            'file': filename,
            'dimension': dimension,
            'confidence_score': details.get('confidence_score', ''),
            'reliability_level': details.get('reliability_level', ''),
            'summary': details.get('summary', ''),
            'total_issues': details.get('total_issues', ''),
            'QA': '',
            'note': ''
        }
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_validation_summary_section(validation_data, filename):
    """Create dimensional analysis summary section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    dimensional_summary = validation_data['validation_report'].get('dimensional_analysis_summary', {})
    rows = []
    
    for dimension, details in dimensional_summary.items():
        row = {
            'file': filename,
            'dimension': dimension,
            'confidence': details.get('confidence', ''),
            'reliability': details.get('reliability', ''),
            'issues_count': details.get('issues_count', ''),
            'QA': '',
            'note': ''
        }
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_validation_critical_section(validation_data, filename):
    """Create critical issues summary section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    critical_issues = validation_data['validation_report'].get('critical_issues_summary', {}).get('issues', [])
    rows = []
    
    for i, issue in enumerate(critical_issues, 1):
        rows.append({
            'file': filename,
            'index': i,
            'critical_issue': issue,
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_markdown_section(markdown_content, filename):
    """Create markdown section dataframe"""
    if not markdown_content:
        return pd.DataFrame()
    
    return pd.DataFrame([{
        'file_name': filename,
        'markdown_content': markdown_content,
        'QA': '',
        'note': ''
    }])

def get_file_list():
    """Get list of all unique file names across all directories"""
    files = set()
    
    # Check each directory for files
    directories = ['results', 'validation_results', 'dataset']
    
    for directory in directories:
        if os.path.exists(directory):
            for filename in os.listdir(directory):
                if filename.endswith('.json'):
                    base_name = filename.replace('.json', '')
                    # Filter out summary_report as it's not an individual file
                    if base_name != 'summary_report':
                        files.add(base_name)
    
    return sorted(list(files))

def generate_detailed_report_for_file(filename):
    """Generate detailed Excel report for a single file"""
    print(f"Processing {filename}...")
    
    # Load data from all sources
    results_data = load_json_file(f'results/{filename}.json')
    validation_data = load_json_file(f'validation_results/{filename}.json')
    markdown_content = load_markdown_file(f'llamaparse_output/{filename}.md')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{filename}_detailed_report"
    
    current_row = 1
    
    # 1. Classification Section
    classification_df = create_classification_section(results_data, filename)
    if not classification_df.empty:
        current_row = add_dataframe_to_worksheet(ws, classification_df, current_row, "CLASSIFICATION SECTION")
    
    # 2. Extraction Results Section
    extraction_df = create_extraction_section(results_data, filename)
    if not extraction_df.empty:
        current_row = add_dataframe_to_worksheet(ws, extraction_df, current_row, "EXTRACTION RESULTS SECTION")
    
    # 3. Issues Section
    issues_df = create_issues_section(results_data, filename)
    if not issues_df.empty:
        current_row = add_dataframe_to_worksheet(ws, issues_df, current_row, "ISSUES SECTION")
    
    # 4. Validation - Overall Assessment
    validation_overall_df = create_validation_overall_section(validation_data, filename)
    if not validation_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_overall_df, current_row, "VALIDATION - OVERALL ASSESSMENT")
    
    # 5. Validation - Dimension Details
    validation_dimensions_df = create_validation_dimensions_section(validation_data, filename)
    if not validation_dimensions_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_dimensions_df, current_row, "VALIDATION - DIMENSION DETAILS")
    
    # 6. Validation - Dimensional Analysis Summary
    validation_summary_df = create_validation_summary_section(validation_data, filename)
    if not validation_summary_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_summary_df, current_row, "VALIDATION - DIMENSIONAL ANALYSIS SUMMARY")
    
    # 7. Validation - Critical Issues
    validation_critical_df = create_validation_critical_section(validation_data, filename)
    if not validation_critical_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_critical_df, current_row, "VALIDATION - CRITICAL ISSUES")
    
    # 8. Markdown Section
    markdown_df = create_markdown_section(markdown_content, filename)
    if not markdown_df.empty:
        current_row = add_dataframe_to_worksheet(ws, markdown_df, current_row, "MARKDOWN SECTION")
    
    # Auto-adjust column widths (fixed version)
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    output_file = f'{filename}_detailed_report.xlsx'
    wb.save(output_file)
    print(f"Generated: {output_file}")
    
    return output_file

def generate_all_detailed_reports():
    """Generate detailed reports for all files"""
    file_list = get_file_list()
    print(f"Found {len(file_list)} files to process: {file_list}")
    
    generated_files = []
    
    for filename in file_list:
        try:
            output_file = generate_detailed_report_for_file(filename)
            generated_files.append(output_file)
        except Exception as e:
            print(f"Error processing {filename}: {e}")
    
    print(f"\nSuccessfully generated {len(generated_files)} detailed reports:")
    for file in generated_files:
        print(f"  - {file}")
    
    return generated_files

if __name__ == "__main__":
    generate_all_detailed_reports()
